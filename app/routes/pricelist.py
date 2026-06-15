"""Price list routes"""
from flask import Blueprint, jsonify, request
from database.models import Material

pricelist_bp = Blueprint('pricelist', __name__, url_prefix='/pricelist')

@pricelist_bp.route('', methods=['GET'])
def get_price_list():
    """Get complete price list"""
    material_type = request.args.get('type')  # 'marble' or 'porcelain'
    
    query = Material.query.filter_by(is_active=True)
    if material_type:
        query = query.filter_by(material_type=material_type)
    
    materials = query.all()
    
    price_list = [{
        'id': m.id,
        'code': m.code,
        'name': m.name,
        'type': m.material_type,
        'description': m.description,
        'color': m.color,
        'finish': m.finish,
        'price_per_sqft': m.price_per_sqft,
        'price_per_meter': m.price_per_meter,
        'stock_quantity': m.stock_quantity,
        'supplier': m.supplier
    } for m in materials]
    
    return jsonify({
        'data': price_list,
        'total': len(price_list),
        'generated_at': __import__('datetime').datetime.utcnow().isoformat()
    })

@pricelist_bp.route('/export', methods=['GET'])
def export_price_list():
    """Export price list as Excel"""
    from flask import send_file
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from datetime import datetime
    
    material_type = request.args.get('type')
    
    query = Material.query.filter_by(is_active=True)
    if material_type:
        query = query.filter_by(material_type=material_type)
    
    materials = query.all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price List"
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    # Header
    ws['A1'] = "INFINITY MARBLE DESIGN - PRICE LIST"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells('A2:H2')
    
    # Column headers
    headers = ['Code', 'Material Name', 'Type', 'Color', 'Finish', 'Price/sqft', 'Price/meter', 'Stock']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    row = 5
    total_value = 0
    for material in materials:
        ws.cell(row=row, column=1).value = material.code
        ws.cell(row=row, column=2).value = material.name
        ws.cell(row=row, column=3).value = material.material_type.upper()
        ws.cell(row=row, column=4).value = material.color
        ws.cell(row=row, column=5).value = material.finish
        
        price_sqft_cell = ws.cell(row=row, column=6)
        price_sqft_cell.value = material.price_per_sqft
        price_sqft_cell.number_format = '$#,##0.00'
        
        price_meter_cell = ws.cell(row=row, column=7)
        price_meter_cell.value = material.price_per_meter
        price_meter_cell.number_format = '$#,##0.00'
        
        stock_cell = ws.cell(row=row, column=8)
        stock_cell.value = material.stock_quantity
        stock_cell.number_format = '0.00'
        
        # Alternate row colors
        if row % 2 == 0:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        row += 1
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'price_list_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'
    )

@pricelist_bp.route('/marble', methods=['GET'])
def get_marble_price_list():
    """Get marble materials price list"""
    materials = Material.query.filter_by(
        material_type='marble',
        is_active=True
    ).all()
    
    return jsonify({
        'type': 'marble',
        'count': len(materials),
        'materials': [{
            'id': m.id,
            'code': m.code,
            'name': m.name,
            'color': m.color,
            'finish': m.finish,
            'price_per_sqft': m.price_per_sqft,
            'origin': m.origin,
            'supplier': m.supplier
        } for m in materials]
    })

@pricelist_bp.route('/porcelain', methods=['GET'])
def get_porcelain_price_list():
    """Get porcelain tiles price list"""
    materials = Material.query.filter_by(
        material_type='porcelain',
        is_active=True
    ).all()
    
    return jsonify({
        'type': 'porcelain',
        'count': len(materials),
        'materials': [{
            'id': m.id,
            'code': m.code,
            'name': m.name,
            'color': m.color,
            'finish': m.finish,
            'price_per_sqft': m.price_per_sqft,
            'origin': m.origin,
            'supplier': m.supplier
        } for m in materials]
    })
